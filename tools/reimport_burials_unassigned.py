import re
import sqlite3
from datetime import date

import openpyxl
from dateutil import parser


def normalize_text(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_possible_lot_number(value):
    if value is None:
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    s = normalize_text(value)
    if s and s.isdigit():
        try:
            return int(s)
        except ValueError:
            return None
    return None


def parse_date_iso(raw):
    s = normalize_text(raw)
    if not s:
        return None
    s = s.replace("\u00a0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if re.fullmatch(r"\d{4}", s):
        return None
    has_year = re.search(r"\b(19|20)\d{2}\b", s) is not None
    if not has_year:
        return None
    try:
        dt = parser.parse(s, fuzzy=True).date()
        return dt.isoformat()
    except Exception:
        return None


def compute_age(dob_iso, dod_iso):
    if not dob_iso or not dod_iso:
        return None
    try:
        y1, m1, d1 = [int(x) for x in dob_iso.split("-")]
        y2, m2, d2 = [int(x) for x in dod_iso.split("-")]
        born = date(y1, m1, d1)
        died = date(y2, m2, d2)
        if died < born:
            return None
        years = died.year - born.year
        if (died.month, died.day) < (born.month, born.day):
            years -= 1
        return years
    except Exception:
        return None


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    check_mark = "\u2705"
    states = {
        1: {"lot": 1, "pending": None, "records": []},
        3: {"lot": 1, "pending": None, "records": []},
    }

    def finalize(col):
        pending = states[col]["pending"]
        if not pending:
            return
        full_name = normalize_text(pending.get("full_name"))
        if not full_name or full_name == check_mark:
            states[col]["pending"] = None
            return
        states[col]["records"].append(pending)
        states[col]["pending"] = None

    for r in range(1, ws.max_row + 1):
        for col in (1, 3):
            cell_val = ws.cell(r, col).value
            if cell_val is None:
                continue

            lot_num = parse_possible_lot_number(cell_val)
            if lot_num is not None:
                finalize(col)
                states[col]["lot"] = lot_num
                continue

            s = normalize_text(cell_val)
            if not s:
                continue
            if s == check_mark:
                continue

            low = s.lower()
            if low.startswith("born:"):
                if states[col]["pending"] is not None:
                    states[col]["pending"]["dob_raw"] = s.split(":", 1)[1].strip()
                continue
            if low.startswith("died:"):
                if states[col]["pending"] is not None:
                    states[col]["pending"]["dod_raw"] = s.split(":", 1)[1].strip()
                continue

            finalize(col)
            states[col]["pending"] = {
                "lot_num": states[col]["lot"],
                "full_name": s,
                "dob_raw": None,
                "dod_raw": None,
            }

    finalize(1)
    finalize(3)

    return states[1]["records"] + states[3]["records"]


def migrate_deceased_records_allow_null_lot(con):
    cur = con.cursor()
    col = cur.execute("PRAGMA table_info(deceased_records)").fetchall()
    lot_col = next((c for c in col if c[1] == "lot_id"), None)
    if not lot_col:
        raise RuntimeError("deceased_records.lot_id column not found")
    notnull = int(lot_col[3] or 0)
    if notnull == 0:
        return False

    con.execute("PRAGMA foreign_keys=OFF")
    con.execute("BEGIN")
    con.execute(
        """
        CREATE TABLE deceased_records_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lot_id INTEGER,
            full_name VARCHAR(255) NOT NULL,
            date_of_birth DATE,
            date_of_death DATE,
            date_of_burial DATE,
            age INTEGER,
            cause_of_death TEXT,
            next_of_kin VARCHAR(255),
            next_of_kin_contact VARCHAR(100),
            remarks TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            layer INTEGER DEFAULT 1,
            deceased_info TEXT,
            is_archived BOOLEAN DEFAULT 0,
            FOREIGN KEY (lot_id) REFERENCES cemetery_lots(id) ON DELETE SET NULL
        )
        """
    )
    con.execute(
        """
        INSERT INTO deceased_records_new
        (id, lot_id, full_name, date_of_birth, date_of_death, date_of_burial, age,
         cause_of_death, next_of_kin, next_of_kin_contact, remarks, created_at, updated_at,
         layer, deceased_info, is_archived)
        SELECT
         id, lot_id, full_name, date_of_birth, date_of_death, date_of_burial, age,
         cause_of_death, next_of_kin, next_of_kin_contact, remarks, created_at, updated_at,
         layer, deceased_info, is_archived
        FROM deceased_records
        """
    )
    con.execute("DROP TABLE deceased_records")
    con.execute("ALTER TABLE deceased_records_new RENAME TO deceased_records")
    con.execute("CREATE INDEX IF NOT EXISTS idx_deceased_lot ON deceased_records(lot_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_deceased_layer ON deceased_records(layer)")
    con.commit()
    con.execute("PRAGMA foreign_keys=ON")
    return True


def record_exists_unassigned(con, full_name, dob_iso, dod_iso):
    cur = con.cursor()
    cur.execute(
        """
        SELECT id
        FROM deceased_records
        WHERE lot_id IS NULL
          AND full_name = ?
          AND COALESCE(date_of_birth, '') = ?
          AND COALESCE(date_of_death, '') = ?
        LIMIT 1
        """,
        (full_name, dob_iso or "", dod_iso or ""),
    )
    return cur.fetchone() is not None


def main():
    xlsx_path = r"assets\burial records.xlsx"
    db_path = r"database\peaceplot.db"

    con = sqlite3.connect(db_path)
    con.execute("PRAGMA foreign_keys=ON")

    migrated = migrate_deceased_records_allow_null_lot(con)

    rows = parse_excel(xlsx_path)

    inserted = 0
    skipped = 0

    con.execute("BEGIN")
    try:
        for row in rows:
            full_name = normalize_text(row.get("full_name")) or "NA"
            dob_raw = normalize_text(row.get("dob_raw"))
            dod_raw = normalize_text(row.get("dod_raw"))

            dob_iso = parse_date_iso(dob_raw)
            dod_iso = parse_date_iso(dod_raw)
            age = compute_age(dob_iso, dod_iso)

            deceased_info = None

            if record_exists_unassigned(con, full_name, dob_iso, dod_iso):
                skipped += 1
                continue

            con.execute(
                """
                INSERT INTO deceased_records
                (lot_id, layer, full_name, date_of_birth, date_of_death, date_of_burial, age,
                 cause_of_death, next_of_kin, next_of_kin_contact, deceased_info, remarks)
                VALUES
                (NULL, NULL, ?, ?, ?, NULL, ?, NULL, NULL, NULL, ?, NULL)
                """,
                (full_name, dob_iso, dod_iso, age, deceased_info),
            )
            inserted += 1

        con.commit()
    except Exception:
        con.rollback()
        raise

    print("migrated_schema", bool(migrated))
    print("parsed_rows", len(rows))
    print("inserted_records", inserted)
    print("skipped_duplicates", skipped)


if __name__ == "__main__":
    main()

